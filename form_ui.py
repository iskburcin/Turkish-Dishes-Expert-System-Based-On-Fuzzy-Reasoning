# sys: handle the app's exit and event loops for managing system level functions
import sys

# pandas manuplates the data; read/write the data on excel/csv in this code
import pandas as pd

# PyQt provides a very rich set of necessary widgets to create the GUI which is highly customizable
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QTabWidget,
    QFormLayout,
    QLineEdit,
    QPushButton,
    QLabel,
    QComboBox,
    QMessageBox,
)

# openpyxl load, read, write and save Excel files here (does not support .xls files, only .xlsx)
from openpyxl import Workbook, load_workbook

from fuzzy_logic import evaluate_dish_from_dataset

EXCEL_FILE = "turkish_dishes.xlsx"
CSV_FILE = "turkish_dishes.csv"

# Also, pandas can read and write Excel files, which can be more efficient and also handle larger datasets better.
# As, openpyxl doesn’t support multi-threaded operations and appends data directly, which might not be ideal for larger datasets (no automatic data integrity checks).


def save_to_excel(data):
    try:
        # Load or (if dousnt exist) create new workbook
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Taste", "Spiciness", "Sweetness", "Texture"])

        # Add the new row
        ws.append(data)
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False


def convert_to_csv():
    # openpyxl or csv module could be used, but pandas provides most efficient way to handle data
    try:
        df = pd.read_excel(EXCEL_FILE)
        df.to_csv(CSV_FILE, index=False)
        print("CSV updated successfully!")
    except Exception as e:
        print(f"Error converting to CSV: {e}")


# Main Window which manages the creation and layout of the user interface.
class DishEvaluator(QMainWindow):
    # The class inherits from QMainWindow, which provides the base for all top-level windows in PyQt
    def __init__(self):  # Constructor
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Turkish Dishes Expert System Based On Fuzzy Reasoning")
        self.setGeometry(800, 200, 400, 400)

        # Central Widget
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        # Tabs
        self.tabs = QTabWidget()
        self.tab_add = QWidget()
        self.tab_check = QWidget()
        self.tab_evaluate = QWidget()
        self.tabs.addTab(self.tab_add, "Add Dish")
        self.tabs.addTab(self.tab_check, "Check Suitability")
        self.tabs.addTab(self.tab_evaluate, "Evaluate Dishes")

        # QVBoxLayout sets the top-level layout for parent. There can be only one tll for a widget
        main_layout = QVBoxLayout(central_widget)
        main_layout.addWidget(self.tabs)

        # Tab 1: Add Dish
        self.setup_add_tab()
        # Tab 2: Check Suitability
        self.setup_check_tab()

        # Tab 3: Evaluate Dishes
        self.setup_evaluate_tab()

        self.apply_styles()

    def setup_add_tab(self):
        layout = QFormLayout()  # organize form elements (labels and input fields)
        self.add_name = QLineEdit()
        self.add_name.setObjectName("dish-name")
        self.add_taste = QLineEdit()
        self.add_spiciness = QLineEdit()
        self.add_sweetness = QLineEdit()
        self.add_texture = QLineEdit()
        self.add_button = QPushButton("Add Dish")

        # connected to the add_dish() function, which handles the dish addition logic
        self.add_button.clicked.connect(self.add_dish)

        layout.addRow("Name:", self.add_name)
        layout.addRow("Taste (0-20):", self.add_taste)
        layout.addRow("Spiciness (0-10):", self.add_spiciness)
        layout.addRow("Sweetness (0-10):", self.add_sweetness)
        layout.addRow("Texture (0-10):", self.add_texture)
        layout.addRow(self.add_button)
        self.tab_add.setLayout(layout)

    def evaluate_dish(self):
        try:
            dish_name = self.dish_selector.currentText()
            logic_choice = self.evaluate_logic_choice.currentIndex() + 1

            # Get membership values
            taste_score, spiciness_score, sweetness_score, texture_score = (
                evaluate_dish_from_dataset(EXCEL_FILE, dish_name, logic_choice)
            )

            # Update individual membership values
            self.taste_value.setText(
                f"Taste Membership Value: {float(taste_score):.2f}"
            )
            self.spiciness_value.setText(
                f"Spiciness Membership Value: {float(spiciness_score):.2f}"
            )
            self.sweetness_value.setText(
                f"Sweetness Membership Value: {float(sweetness_score):.2f}"
            )
            self.texture_value.setText(
                f"Texture Membership Value: {float(texture_score):.2f}"
            )

            # Calculate suitability using fuzzy inference
            suitability = self.fuzzy_inference(
                taste_score,
                spiciness_score,
                sweetness_score,
                texture_score,
                logic_choice,
            )
            self.evaluate_result.setText(f"Suitability Score: {suitability:.2f}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to evaluate dish: {e}")

    def setup_evaluate_tab(self):
        layout = QVBoxLayout()

        # Dish Selection Dropdown
        self.dish_selector = QComboBox()
        self.load_dishes()

        # Membership Function Dropdown
        self.evaluate_logic_choice = QComboBox()
        self.evaluate_logic_choice.addItem("Logic 1: Triangular")
        self.evaluate_logic_choice.addItem("Logic 2: Trapezoidal")
        self.evaluate_logic_choice.addItem("Logic 3: Gaussian")

        # Evaluate Button
        self.evaluate_button = QPushButton("Evaluate Dish")
        self.evaluate_button.clicked.connect(self.evaluate_dish)

        # Labels for individual membership results
        self.taste_value = QLabel("Taste Membership Value: ")
        self.spiciness_value = QLabel("Spiciness Membership Value: ")
        self.sweetness_value = QLabel("Sweetness Membership Value: ")
        self.texture_value = QLabel("Texture Membership Value: ")
        self.evaluate_result = QLabel("Suitability Score: ")

        # Add Widgets to Layout
        x = QHBoxLayout()
        x.addWidget(QLabel("Select Dish:"))
        x.addWidget(self.dish_selector)

        y = QHBoxLayout()
        y.addWidget(QLabel("Select Membership Function:"))
        y.addWidget(self.evaluate_logic_choice)

        layout.addLayout(x)
        layout.addLayout(y)
        layout.addWidget(self.evaluate_button)
        layout.addWidget(self.taste_value)
        layout.addWidget(self.spiciness_value)
        layout.addWidget(self.sweetness_value)
        layout.addWidget(self.texture_value)
        layout.addWidget(self.evaluate_result)

        self.tab_evaluate.setLayout(layout)

    def load_dishes(self):
        """Load dishes from the dataset into the dropdown."""
        try:
            df = (
                pd.read_excel(EXCEL_FILE)
                if EXCEL_FILE.endswith(".xlsx")
                else pd.read_csv(CSV_FILE)
            )
            self.dish_selector.clear()
            self.dish_selector.addItems(df["Name"].tolist())
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load dishes: {e}")

    def setup_check_tab(self):
        layout = QFormLayout()

        # Input fields for all four characteristics
        self.check_taste = QLineEdit()
        self.check_spiciness = QLineEdit()
        self.check_sweetness = QLineEdit()
        self.check_texture = QLineEdit()
        self.check_taste.setProperty("mandatoryField", True)
        self.check_spiciness.setProperty("mandatoryField", True)
        self.check_sweetness.setProperty("mandatoryField", True)
        self.check_texture.setProperty("mandatoryField", True)

        # Dropdown for logic
        self.logic_choice = QComboBox()
        self.logic_choice.addItem("Logic 1: Triangular")
        self.logic_choice.addItem("Logic 2: Trapezoidal")
        self.logic_choice.addItem("Logic 3: Gaussian")

        # Button and result label
        self.check_result = QLabel("Result: ")
        self.check_button = QPushButton("Check Suitability")
        self.check_button.clicked.connect(self.check_suitability)

        # Add widgets to layout
        layout.addRow("Taste (0-20):", self.check_taste)
        layout.addRow("Spiciness (0-10):", self.check_spiciness)
        layout.addRow("Sweetness (0-10):", self.check_sweetness)
        layout.addRow("Texture (0-10):", self.check_texture)
        layout.addRow("Select Logic:", self.logic_choice)
        layout.addRow(self.check_button)
        layout.addRow(self.check_result)
        self.tab_check.setLayout(layout)

    def add_dish(self):
        try:
            # pythonda dynamic allocation olduğu için stirng algılanan texti floata parselanmalı
            name = self.add_name.text()
            taste = float(self.add_taste.text())
            spiciness = float(self.add_spiciness.text())
            sweetness = float(self.add_sweetness.text())
            texture = float(self.add_texture.text())

            if not (
                0 <= taste <= 20
                and 0 <= spiciness <= 10
                and 0 <= sweetness <= 10
                and 0 <= texture <= 10
            ):
                raise ValueError

            # Save data
            if save_to_excel([name, taste, spiciness, sweetness, texture]):
                QMessageBox.information(self, "Success", "Dish added successfully!")
                self.add_name.clear()
                self.add_taste.clear()
                self.add_spiciness.clear()
                self.add_sweetness.clear()
                self.add_texture.clear()

                # Update CSV after 10 dishes
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
                if len(ws["A"]) % 10 == 0:  # Check if 10 dishes are added
                    convert_to_csv()
            else:
                QMessageBox.critical(self, "Error", "Failed to add dish.")
        except ValueError:
            QMessageBox.critical(
                self, "Error", "Enter valid numbers within the specified ranges!"
            )

    def check_suitability(self):
        from fuzzy_logic import predict_suitability

        try:
            # Get inputs
            taste = float(self.check_taste.text())
            spiciness = float(self.check_spiciness.text())
            sweetness = float(self.check_sweetness.text())
            texture = float(self.check_texture.text())
            logic_choice = self.logic_choice.currentIndex() + 1

            # Evaluate suitability
            result = predict_suitability(
                taste, spiciness, sweetness, texture, logic_choice
            )
            self.check_result.setText(result)
        except ValueError:
            QMessageBox.critical(
                self, "Error", "Enter valid numbers within the specified ranges!"
            )

    def fuzzy_inference(
        self, taste_score, spiciness_score, sweetness_score, texture_score, logic_choice
    ):
        from fuzzy_logic import predict_suitability

        # Get inputs
        taste = float(taste_score)
        spiciness = float(spiciness_score)
        sweetness = float(sweetness_score)
        texture = float(texture_score)

        # Evaluate suitability
        result = predict_suitability(
            taste, spiciness, sweetness, texture, logic_choice, ispredict=True
        )
        return float(result)

    def apply_styles(self):
        # Applies custom CSS-like styling using Qt Style Sheets (QSS)
        # Not as efficient as hard-coding the styles if advanced styling is needed
        qss = """
        QWidget {
            font-family: Arial, sans-serif;
            font-size: 14px;
            color: white;
        }
        QPushButton {
            background-color: #0078D7;
            border-radius: 5px;
            padding: 5px;
        }
        QPushButton:hover {
            background-color: #005A9E;
        }
        QLineEdit {
            border: 1px solid #CCCCCC;
            border-radius: 5px;
            padding: 4px;
        }
        QLineEdit#dish-name{
            background-color: grey
        }
        QLabel {
            color: white;
        }
        *[mandatoryField="True"] {
            QLineEdit {
                border: 1px solid darkred;
            }
        }
        """
        self.setStyleSheet(qss)


# Main function
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DishEvaluator()
    window.show()
    sys.exit(app.exec())
